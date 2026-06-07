import openpyxl

file_path = r"e:\1.workspace\7.other\17.charge_demo\charge_demo\05-Reference(official)\requirement_20260518.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"Workbook loaded. Sheet names: {wb.sheetnames}")
print("=" * 100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 100}")
    print(f"Sheet: [{sheet_name}]")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"{'=' * 100}")

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False), start=1):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            row_data.append(f"[{col_idx}] {repr(val)}")
        print(f"Row {row_idx:4d}: {' | '.join(row_data)}")

wb.close()
print("\nDone.")
