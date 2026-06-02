import openpyxl
import fitz

# Read Excel
wb = openpyxl.load_workbook(r'e:\1.workspace\7.other\17.charge_demo\charge_demo\05-Reference(official)\requirement_20260518.xlsx')
ws = wb.active
print('=== EXCEL REQUIREMENT ===')
print('Sheet:', ws.title)
print('Rows:', ws.max_row, 'Cols:', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    vals = [str(cell.value) if cell.value is not None else '' for cell in row]
    print(' | '.join(vals))

print()
print('=== PDF SCHEMATIC ===')
doc = fitz.open(r'e:\1.workspace\7.other\17.charge_demo\charge_demo\02-sch\L1211 TOP V2.0(1).pdf')
for i, page in enumerate(doc):
    print(f'--- Page {i+1} ---')
    text = page.get_text()
    print(text)
doc.close()