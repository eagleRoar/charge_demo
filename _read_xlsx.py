import openpyxl
wb = openpyxl.load_workbook(r'e:\1.workspace\7.other\17.charge_demo\charge_demo\05-Reference(official)\requirement_20260518.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== Sheet: {name} ===")
    for row in ws.iter_rows():
        print('\t'.join([str(c.value or '') for c in row]))
